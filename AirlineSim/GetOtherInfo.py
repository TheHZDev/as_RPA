import json
import sqlite3
from threading import Thread
from time import sleep
from urllib.request import getproxies

import requests
from bs4 import BeautifulSoup
from bs4.element import Tag as bs4_Tag

from PublicCode import CommonHTMLParser, GetClearHTML

max_thread_workers = 5
basic_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:93.0) Gecko/20100101 Firefox/93.0'}
LocalProxier = {'http': '', 'https': ''}
LocalProxier.update(getproxies())
try:
    from local_debug import flag_Debug

    Debug_Allow_TLS_Verify = not flag_Debug
except:
    Debug_Allow_TLS_Verify = True


def retry_request(url: str, timeout: int = 60, retry_cooldown: int = 5, retry_times: int = 3, data=None):
    """
    规范化网络连接尝试规程，以解决HTTPS连接慢及服务器垃圾问题
    :param url: 要连接的目标URL，以GET的方式连接
    :param timeout: 超时时间
    :param retry_cooldown: 重试前的冷却时间
    :param retry_times: 最大尝试次数
    :param data: 要POST到服务端的数据
    """
    if timeout < 1:
        timeout = 60
    retry_cooldown = max(5, retry_cooldown)
    retry_times = max(3, retry_times)
    for i in range(retry_times):
        try:
            if data is None:
                return requests.get(url, timeout=timeout, headers=basic_header, proxies=LocalProxier,
                                    verify=Debug_Allow_TLS_Verify)
            else:
                return requests.post(url, data=data, timeout=timeout, headers=basic_header, proxies=LocalProxier,
                                     verify=Debug_Allow_TLS_Verify)
        except requests.exceptions.ConnectTimeout:
            print('连接出错，将在 %d 秒后再次重试。' % retry_cooldown)
            sleep(retry_cooldown)
    return requests.Response()


def retry_request_Session(session: requests.Session, url: str, timeout: int = 60, retry_cooldown: int = 5,
                          retry_times: int = 3, data=None):
    if timeout < 1:
        timeout = 60
    retry_cooldown = max(5, retry_cooldown)
    retry_times = max(3, retry_times)
    for i in range(retry_times):
        try:
            if data is None:
                return session.get(url, timeout=timeout, proxies=LocalProxier, verify=Debug_Allow_TLS_Verify)
            else:
                return session.post(url, data=data, timeout=timeout, proxies=LocalProxier,
                                    verify=Debug_Allow_TLS_Verify)
        except:
            print('连接出错，将在 %d 秒后再次重试。' % retry_cooldown)
            sleep(retry_cooldown)
    return requests.Response()


class ConfigManager:

    @staticmethod
    def CreateSQL():
        create_sql = """
        CREATE TABLE IF NOT EXISTS ConfigManager(
            Module TEXT PRIMARY KEY,
            JSON_Config TEXT
        );
        """
        return create_sql

    @staticmethod
    def SaveConfig(DBPath: str, ModuleName: str, Config: dict) -> bool:
        t_sql = sqlite3.connect(DBPath)
        flag_success = True
        create_sql = """
        CREATE TABLE IF NOT EXISTS ConfigManager(
            Module TEXT PRIMARY KEY,
            JSON_Config TEXT
        );
        """
        t_sql.execute(create_sql)  # 出于便携化的要求
        try:
            Config = json.dumps(Config)
            select_sql = "SELECT COUNT(*) FROM ConfigManager WHERE Module = ?;"
            insert_sql = "INSERT INTO ConfigManager VALUES(?, ?);"
            update_sql = "UPDATE ConfigManager SET JSON_Config = ? WHERE Module = ?;"
            if t_sql.execute(select_sql, (ModuleName,)).fetchone()[0] == 1:
                t_sql.execute(update_sql, (Config, ModuleName))
            else:
                t_sql.execute(insert_sql, (ModuleName, Config))
            t_sql.commit()
        except:
            flag_success = False
        finally:
            t_sql.close()
            return flag_success

    @staticmethod
    def LoadConfig(DBPath: str, ModuleName: str) -> dict:
        Config = {}
        t_sql = sqlite3.connect(DBPath)
        try:
            select_sql = "SELECT JSON_Config FROM ConfigManager WHERE Module = ?;"
            Config = json.loads(t_sql.execute(select_sql, (ModuleName,)).fetchone()[0])
        finally:
            t_sql.close()
            return Config


class CalcAirplaneProperty:
    login_UserName = None
    login_Passwd = None
    # 缓存信息块
    cache_CountryIndex = []
    cache_AirCompanyURL = []
    cache_CabinAnalyze = {}
    flag_price_ok = False

    def __init__(self, ServerName: str, UserName: str = '', Passwd: str = ''):
        from LoginAirlineSim import ServerMap, getBaseURL
        if ServerName not in ServerMap.keys():
            raise Exception('无效的服务器名称。')
        self.baseURL = getBaseURL(ServerName)
        self.baseURL_AirCompany = self.baseURL + '/action/info/'
        self.ServerName = ServerName
        self.baseDB = ServerName + '.sqlite'
        if isinstance(UserName, str) and isinstance(Passwd, str) and len(UserName) * len(Passwd) > 1:
            self.login_Passwd = Passwd
            self.login_UserName = UserName
        self.DB_Init()

    def DB_Init(self):
        t1 = sqlite3.connect(self.baseDB)
        create_sql = """
        CREATE TABLE IF NOT EXISTS Fleets(
            AirCompany TEXT,
            AirType TEXT,
            FirstCabin INTEGER,
            BusinessCabin INTEGER,
            EconomyCabin INTEGER,
            IsRented INTEGER,
            IsDelivering INTEGER
        );
        """
        # AirCompany - 航空公司全称
        # AirType - 航机类型，即型号
        # FirstCabin - 头等舱的舱位数量
        # BusinessCabin - 商务舱的舱位数量
        # EconomyCabin - 经济舱的舱位数量
        # IsRented - 是否为租赁的航机
        # IsDelivering - 是否尚未交机
        t1.execute(create_sql)
        clear_sql = "DELETE FROM Fleets;"
        t1.execute(clear_sql)
        create_sql = """
        CREATE TABLE IF NOT EXISTS AirCompanyMap(
            AirCompany TEXT,
            ParentAirCompany TEXT
        );
        """
        # AirCompany - 子公司名称
        # ParentAirCompany - 子公司所属的母公司的名称
        t1.execute(create_sql)
        clear_sql = "DELETE FROM AirCompanyMap;"
        t1.execute(clear_sql)
        create_sql = """
        CREATE TABLE IF NOT EXISTS AirplaneInfo(
            AirType TEXT,
            MaxPassenger INTEGER,
            MaxCargo INTEGER,
            MinFlightLength INTEGER,
            MaxFlightLength INTEGER,
            Speed INTEGER,
            MinDepartureRunway INTEGER,
            MaxDepartureRunway INTEGER,
            MinArriveRunway INTEGER,
            MaxArriveRunway INTEGER,
            Price INTEGER
        );
        """
        # AirType - 航机型号
        # MaxPassenger - 最大乘客量
        # MaxCargo - 最大载货量，单位是千克
        # MinFlightLength - 最小航程，单位是千米
        # MaxFlightLength - 最大航程，单位是千米
        # Speed - 巡航速度，单位是千米每小时
        # MinDepartureRunway - 最短降落跑道长度，单位是米
        # MaxDepartureRunway - 最长降落跑道长度，单位是米
        # MinArriveRunway - 最短起飞跑道长度，单位是米
        # MaxArriveRunway - 最长起飞跑道长度，单位是米
        # Price - 航机的购买价格，租赁保证金取该数值的二十分之一
        t1.execute(create_sql)
        t1.commit()

    def getAirplaneInfoIndex(self):
        """获取所有的机队信息，还包括后面可以使用的各家航空公司的URL和航机的URL"""
        first_url = self.baseURL + '/action/info/fleets'

        def Recursion_GetAllCountryID(root: bs4_Tag):
            if root.name == 'select' and root.attrs.get('id', '') == 'country-select':
                prefix_url = first_url + '?%s=' % root.attrs.get('name')
                for t_unit in root.contents:
                    if isinstance(t_unit, bs4_Tag) and t_unit.name == 'option':
                        self.cache_CountryIndex.append(prefix_url + t_unit.attrs.get('value'))
                return
            for t_unit in root.children:
                if len(self.cache_CountryIndex) > 0:
                    return
                if isinstance(t_unit, bs4_Tag):
                    Recursion_GetAllCountryID(t_unit)

        for unit in GetClearHTML(retry_request(first_url)).children:
            if len(self.cache_CountryIndex) > 0:
                break
            if isinstance(unit, bs4_Tag):
                Recursion_GetAllCountryID(unit)
        self.callback_outputLog('抓取国家索引完成，第一阶段结束。')
        for i in range(max_thread_workers):
            Thread(target=self.thread_getAirplaneInfo).start()

    def getAirCompanyInfoIndex(self):
        """获取所有的航空公司信息，包括该航空公司是否属于任何一家公司的母公司"""
        if len(self.cache_AirCompanyURL) > 0:
            for i in range(max_thread_workers):
                Thread(target=self.thread_getAirCompanyInfo).start()
        else:
            self.callback_outputLog('没有可获取的航空公司信息。')

    def thread_getAirplaneInfo(self):
        if len(self.cache_CountryIndex) == 0:
            return
        target_url = self.cache_CountryIndex.pop()
        result_text = ['']

        def Recursion_GetAirplaneInfo(root: bs4_Tag):
            if root.name == 'table':
                t_sql = sqlite3.connect(self.baseDB)
                insert_sql = "INSERT INTO Fleets VALUES(?,?,?,?,?,?,?);"
                for t_unit in root.children:
                    if isinstance(t_unit, bs4_Tag) and t_unit.name == 'tbody':
                        AirCompanyName: str = t_unit.contents[0].contents[0].getText()  # 解析企业数据
                        pre_AirCompany_URL: str = t_unit.contents[1].contents[0].contents[3].attrs.get('href')
                        self.cache_AirCompanyURL.append(self.baseURL_AirCompany + pre_AirCompany_URL)
                        for line in t_unit.contents[3:len(t_unit.contents) - 1]:
                            isRent = 0
                            isDelivering = 0
                            AirType: str = line.contents[1].contents[0].getText()
                            seatData = line.contents[4].getText()
                            if seatData in self.cache_CabinAnalyze.keys():
                                seatModel = self.cache_CabinAnalyze.get(seatData)  # 取缓存的已解析数据
                            else:
                                seatModel = self.AnalyzeCabinModel(seatData)
                                self.cache_CabinAnalyze[seatData] = seatModel  # 大家都来帮一把
                            comment: str = line.contents[5].getText()
                            if 'lsf' in comment:
                                isRent = 1  # 租用的，非自有
                            if 'not yet delivered' in comment:
                                isDelivering = 1  # 尚未交机
                            t_sql.execute(insert_sql, (AirCompanyName, AirType, seatModel[0], seatModel[1],
                                                       seatModel[2], isRent, isDelivering))
                        t_sql.commit()
                t_sql.close()
                return
            elif root.name == 'h3':
                result_text[0] = root.getText()
                return
            for t_unit in root.children:
                if isinstance(t_unit, bs4_Tag):
                    Recursion_GetAirplaneInfo(t_unit)

        for unit in GetClearHTML(retry_request(target_url)).children:
            if isinstance(unit, bs4_Tag):
                Recursion_GetAirplaneInfo(unit)
        self.callback_outputLog('对国家 %s 的信息抓取完成。' % result_text[0])
        Thread(target=self.thread_getAirplaneInfo).start()  # 虚设的线程池

    def thread_getAirCompanyInfo(self):
        if len(self.cache_AirCompanyURL) == 0:
            return
        target_url = self.cache_AirCompanyURL.pop()
        result_text = ['']

        def Recursion_GetAirCompanyInfo(root: bs4_Tag):
            if root.name == 'td' and root.getText() == 'Parent company':
                if root.parent.contents[1].contents[0].getText() != 'Enterprise is a holding':
                    # 企业非控股公司，即，有一个母公司
                    ParentCompany = root.parent.contents[1].contents[0].getText()
                    CompanyName = root.parent.parent.contents[0].contents[1].contents[0].getText()
                    t_sql = sqlite3.connect(self.baseDB)
                    insert_sql = "INSERT INTO AirCompanyMap VALUES(?,?);"
                    try:
                        t_sql.execute(insert_sql, (CompanyName, ParentCompany))
                        t_sql.commit()
                    finally:
                        t_sql.close()
                    return
            elif root.name == 'td' and root.getText() == 'Name':
                result_text[0] = root.parent.contents[1].contents[0].getText()
                return
            for t_unit in root.children:
                if isinstance(t_unit, bs4_Tag):
                    Recursion_GetAirCompanyInfo(t_unit)

        for unit in GetClearHTML(retry_request(target_url)).children:
            if isinstance(unit, bs4_Tag):
                Recursion_GetAirCompanyInfo(unit)
        self.callback_outputLog('对公司 %s 的信息抓取完成。' % result_text[0])
        Thread(target=self.thread_getAirCompanyInfo).start()

    def thread_getAirplanePrice(self):
        t_sql = sqlite3.connect(self.baseDB)
        select_sql = "SELECT COUNT(*) FROM AirplaneInfo;"
        if t_sql.execute(select_sql).fetchone()[0] > 1:
            # 不再重复抓取数据
            t_sql.close()
            return
        if isinstance(self.login_Passwd, str) and isinstance(self.login_UserName, str):
            from LoginAirlineSim import LoginAirlineSim
            logonSession = LoginAirlineSim(self.ServerName, self.login_UserName, self.login_Passwd)
            first_url = self.baseURL + '/app/aircraft/manufacturers'  # 从制造商页面爬取数据
            second_url_list = {}

            def Recursion_GetAllAirplaneFamilyInfo(root: bs4_Tag):
                if root.name == 'a' and root.attrs.get('href', '').startswith(
                        '../../action/enterprise/aircraftsFamily?id='):
                    second_url_list[root.getText()] = self.baseURL + root.attrs.get('href').replace('../..', '')
                    return
                for t_unit in root.children:
                    if isinstance(t_unit, bs4_Tag):
                        Recursion_GetAllAirplaneFamilyInfo(t_unit)

            for unit in GetClearHTML(logonSession.get(first_url)).children:
                if isinstance(unit, bs4_Tag):
                    Recursion_GetAllAirplaneFamilyInfo(unit)
            # 抓取了所有航班家族的数据
            insert_sql = "INSERT INTO AirplaneInfo VALUES(?,?,?,?,?,?,?,?,?,?,?);"

            def Recursion_GetSingleFamilyAirplaneInfo(root: bs4_Tag):
                if root.name == 'a' and root.attrs.get('href', '').startswith('aircraftsType?id='):
                    AirType = root.getText().strip()
                    row_root = root.parent.parent
                    MaxPassenger = int(row_root.contents[1].getText().strip())
                    MaxCargo = int(row_root.contents[2].getText().replace(',', '').replace('kg', '').strip())
                    AirRange = row_root.contents[3].getText().replace('km', '').replace(',', '').strip().split('-')
                    MinFlightLength = int(AirRange[0])
                    MaxFlightLength = int(AirRange[1])
                    Speed = int(row_root.contents[4].getText().replace('km/h', '').strip())
                    AirRunway = row_root.contents[5].getText().replace('m', '').replace(',', '').strip().split('-')
                    MinDepartRunway = int(AirRunway[0])
                    MaxDepartRunway = int(AirRunway[1])
                    AirRunway = row_root.contents[6].getText().replace('m', '').replace(',', '').strip().split('-')
                    MinArriveRunway = int(AirRunway[0])
                    MaxArriveRunway = int(AirRunway[1])
                    Price = int(row_root.contents[7].getText().replace('AS$', '').replace(',', '').strip())
                    t_sql.execute(insert_sql, (AirType, MaxPassenger, MaxCargo, MinFlightLength, MaxFlightLength,
                                               Speed, MinDepartRunway, MaxDepartRunway, MinArriveRunway,
                                               MaxArriveRunway, Price))
                for t_unit in root.children:
                    if isinstance(t_unit, bs4_Tag):
                        Recursion_GetSingleFamilyAirplaneInfo(t_unit)

            for line in second_url_list.keys():
                for unit in GetClearHTML(logonSession.get(second_url_list.get(line))).children:
                    if isinstance(unit, bs4_Tag):
                        Recursion_GetSingleFamilyAirplaneInfo(unit)
                t_sql.commit()
                self.callback_outputLog('已完成对航机 %s 家族的爬取。' % line)
            t_sql.close()
            from LoginAirlineSim import LogoutAirlineSim
            LogoutAirlineSim(logonSession)
        self.flag_price_ok = True

    @staticmethod
    def AnalyzeCabinModel(CabinStr: str):
        if len(CabinStr) == 0:
            return 0, 0, 0
        FirstCabin = 0
        BusinessCabin = 0
        EconomyCabin = 0
        if CabinStr.startswith('F'):
            if 'C' in CabinStr and 'Y' in CabinStr:
                t1 = CabinStr[1:].split('C')
                FirstCabin = int(t1[0])
                t1 = t1[1].split('Y')
                BusinessCabin = int(t1[0])
                EconomyCabin = int(t1[1])
            elif 'C' in CabinStr:
                t1 = CabinStr[1:].split('C')
                FirstCabin = int(t1[0])
                BusinessCabin = int(t1[1])
            elif 'Y' in CabinStr:
                t1 = CabinStr[1:].split('Y')
                FirstCabin = int(t1[0])
                EconomyCabin = int(t1[1])
            else:
                FirstCabin = int(CabinStr[1:])
        elif CabinStr.startswith('C'):
            if 'Y' in CabinStr:
                t1 = CabinStr[1:].split('Y')
                BusinessCabin = int(t1[0])
                EconomyCabin = int(t1[1])
            else:
                BusinessCabin = int(CabinStr[1:])
        elif CabinStr.startswith('Y'):
            EconomyCabin = int(CabinStr[1:])
        return FirstCabin, BusinessCabin, EconomyCabin

    def callback_outputLog(self, logStr: str):
        # 预留日志输出接口
        print(logStr)

    def CalcBalanceSheet(self, SeatPrice=(4400, 1900, 500), MergeSubCompany: bool = True,
                         DescSorted: bool = True):
        """
        计算各航空公司的资产负债表\n
        :param SeatPrice: 席位价格表组，从左至右为头等舱、商务舱、经济舱
        :param MergeSubCompany: 是否归并子公司，归并后子公司资产作为母公司的一部分进行计算
        :param DescSorted: 是否降序排列，即总资产最高的将排在前面
        :return: 资产负债表的list
        """
        t_sql = sqlite3.connect(self.baseDB)
        select_sql = "SELECT DISTINCT AirCompany FROM Fleets;"
        result_dict = {}
        for AirCompany in t_sql.execute(select_sql).fetchall():
            result_dict[AirCompany[0]] = [0, 0]  # 211218改动：添加了计算已购航机资产的功能
        if len(result_dict) == 0:
            return []
        cache_AirPriceMap = {}  # 存储航机价格表
        select_sql = "SELECT AirType, Price FROM AirplaneInfo WHERE AirType in (SELECT DISTINCT AirType " \
                     "FROM Fleets);"
        for line in t_sql.execute(select_sql).fetchall():
            cache_AirPriceMap[line[0]] = line[1]
        if len(cache_AirPriceMap) == 0:
            raise Exception('未获取到航机价格信息，无法启动计算！')
        self.callback_outputLog('计算资产中......数据缓存建立完毕！')
        # 初始化完毕
        select_sql = "SELECT AirType, FirstCabin, BusinessCabin, EconomyCabin, IsRented FROM Fleets " \
                     "WHERE AirCompany = ?;"
        for AirCompany in result_dict.keys():
            for line in t_sql.execute(select_sql, (AirCompany,)).fetchall():
                if line[4] == 1:
                    result_dict[AirCompany][0] += cache_AirPriceMap.get(line[0]) / 20
                else:
                    result_dict[AirCompany][0] += cache_AirPriceMap.get(line[0])
                    result_dict[AirCompany][1] += cache_AirPriceMap.get(line[0]) + (
                            line[1] * SeatPrice[0] + line[2] * SeatPrice[1] + line[3] * SeatPrice[2])
                    # 已购飞机的资产另外列出，但原有合并计算代码保持不变
                result_dict[AirCompany][0] += line[1] * SeatPrice[0] + line[2] * SeatPrice[1] + line[3] * SeatPrice[2]

        self.callback_outputLog('计算资产中......所有航机及座位资产计算完成！')
        # 初始计算完成
        if MergeSubCompany:
            select_sql = "SELECT AirCompany, ParentAirCompany FROM AirCompanyMap;"
            cache_SubCompany = {}
            for line in t_sql.execute(select_sql).fetchall():
                cache_SubCompany[line[0]] = line[1]  # 子公司: 对应的母公司
            # 建立子公司映射表缓存
            t_list = []
            for line in cache_SubCompany.keys():
                if line not in result_dict.keys():
                    t_list.append(line)
            for line in t_list:
                # 底下并没有任何航机的子公司不参与后续的归并计算（防止计算出错）
                cache_SubCompany.pop(line)
            flag_continue_merge = True
            while flag_continue_merge:
                for line in cache_SubCompany.keys():
                    if result_dict.get(line)[0] > 0:
                        if cache_SubCompany.get(line) not in result_dict.keys():
                            # 考虑到某些企业并没有实际运营航空公司而是交由子公司运营的情况
                            result_dict[cache_SubCompany.get(line)] = [0, 0]
                        result_dict[cache_SubCompany[line]][0] += result_dict.get(line)[0]
                        result_dict[line][0] = 0
                        result_dict[cache_SubCompany[line]][1] += result_dict.get(line)[1]
                        result_dict[line][1] = 0
                # 循环归并，相当于冒泡排序
                flag_continue_merge = False
                for line in cache_SubCompany.keys():
                    if result_dict.get(line)[0] > 0:
                        # 当满足所有子公司资产数额都为0的时候，归并计算结束
                        flag_continue_merge = True
                        break
            for line in cache_SubCompany.keys():
                result_dict.pop(line)  # 删除资产为0的子公司
            self.callback_outputLog('计算资产中......子公司归并计算完毕！')
        t_sql.close()
        result_list = []
        for AirCompany in result_dict.keys():
            result_list.append((AirCompany, result_dict.get(AirCompany)[0], result_dict.get(AirCompany)[1]))

        def _cmp(x, y):
            if x[1] > y[1]:
                return 1
            else:
                return -1

        from functools import cmp_to_key
        result_list.sort(key=cmp_to_key(_cmp), reverse=DescSorted)
        self.callback_outputLog('计算资产中......数据已按照资产数额排列完成！')
        return result_list

    # 新增导出函数（控制台，难看死了）
    def OutputPropertyToExcel(self, CalcData: list, FilePath: str = ''):
        """将资产数据导出至Excel文件"""
        try:
            import openpyxl
            tWorkBook = openpyxl.Workbook()
            tTable = tWorkBook.active
            tTable.title = self.ServerName
            tTable.append(('排名', '企业名称', '资产数额/K AS$', '已购飞机资产额/K AS$'))
            order_int = 1
            for line in CalcData:
                tTable.append((str(order_int), line[0], str(line[1] / 1000), str(line[2] / 1000)))
                order_int += 1
            if FilePath == '' or not isinstance(FilePath, str):
                from datetime import datetime
                FilePath = datetime.now().strftime('%Y%m%d.xlsx')
                self.callback_outputLog('将导出到%s。' % FilePath)
            if not FilePath.endswith('.xlsx'):
                FilePath += '.xlsx'
            tWorkBook.save(FilePath)
        except ModuleNotFoundError:
            self.callback_outputLog('请先安装openpyxl模块，具体为"pip install openpyxl"！')
        except:
            self.callback_outputLog('未知错误，导出失败！')

    def OutputPropertyToHTML(self, CalcData: list, FilePath: str = ''):
        """导出资产数据到HTML页面"""
        from os.path import isfile
        from os import getcwd, sep
        templateHTML_path = getcwd() + sep + 'TemplateHTML' + sep + 'GetOtherInfo_OutputHTML_Template.html'
        if not isfile(templateHTML_path):
            raise FileNotFoundError('导出失败！请重新下载HTML模板文件！')
        # 替换对应行数据（索引53）
        templateHTML = open(templateHTML_path, 'r', encoding='UTF-8')
        list_html = templateHTML.read().splitlines()
        templateHTML.close()
        pre_html_load = ''.join(
            ["AddNewTableLine('%s', '%.2f', '%.2f');" % (i[0], i[1] / 1000, i[2] / 1000) for i in CalcData])
        list_html[53] = pre_html_load  # 这里实际上是替换了window.onload函数，实现表格数据添加
        if FilePath == '' or not isinstance(FilePath, str):
            from datetime import datetime
            FilePath = datetime.now().strftime('%Y%m%d.html')
            self.callback_outputLog('将导出到%s。' % FilePath)
        if not FilePath.endswith('.html'):
            FilePath += '.html'
        tFile = open(FilePath, 'w', encoding='UTF-8')
        tFile.write('\r\n'.join(list_html))
        tFile.close()


class GetAirportInfo:
    flag_finish = []

    def __init__(self, ServerName: str):
        """
        一个用来抓取机场信息的小工具，没有任何其他的计算功能
        :param ServerName: 服务器名称
        """
        from LoginAirlineSim import ServerMap, getBaseURL
        from urllib.parse import urlparse
        if ServerName not in ServerMap.keys():
            raise Exception('无效的服务器名称。')
        self.baseURL = 'https://' + urlparse(getBaseURL(ServerName)).netloc + '/app/info/airports/'
        self.errorURL = '/app/wicket/page'  # URL后缀
        self.DBPath = ServerName + '.sqlite'
        self.DB_Init()
        self.get_header = basic_header.copy()
        self.get_header['Accept-Language'] = 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2'
        # 判断机场信息以及有无更新
        t_sql = sqlite3.connect(self.DBPath)
        t1 = t_sql.execute('SELECT COUNT(*) FROM AirportInfo;').fetchone()[0]
        t_sql.close()
        if t1 == 0 or self.DetectAirportUpdate_Basic():
            self.DB_TRUNCATE()
            for i in range(1, max_thread_workers + 1):
                Thread(target=self.thread_getAirportInfo, args=(i,)).start()
        else:
            print('未检测到更新，且已爬取数据。若上次未能正常退出，请执行DB_TRUNCATE函数以初始化。')

    def DB_Init(self):
        create_sql = """
        CREATE TABLE IF NOT EXISTS AirportInfo(
            AirportID INTEGER,
            AirportName TEXT,
            Time_Zone INTEGER,
            IATA TEXT,
            ICAO TEXT,
            Country TEXT,
            Region TEXT,
            Continent TEXT,
            Runway INTEGER,
            Airport_Size TEXT,
            Slots_per_five_minutes INTEGER,
            Slot_Availability FLOAT,
            Min_transfer_time INTEGER,
            Nighttime_ban INTEGER,
            Noise_restrictions INTEGER,
            Passengers INTEGER,
            Cargo INTEGER
        );
        """
        # AirportID - 机场的登记ID，这个是为了日后定制更新机场信息而设计
        # Time_Zone - 时区，以UTC时间作为标定
        # IATA - 国际航空运输协会（International Air Transport Association）代码
        # ICAO - 国际民用航空组织（International Civil Aviation Organization）代码
        # Country - 机场所在国名称，可惜是英文的
        # Region - 机场在所在国的区域，可惜是英文的
        # Continent - 国家所在的洲
        # Runway - 跑道长度，单位是米
        # Airport_Size - 机场大小
        # Slots_per_five_minutes - 每5分钟的时间带
        # Slot_Availability - 时间带可用百分比，应该是指示机场的繁忙程度
        # Min_transfer_time - 最小转机时间，如果这个值是-1，则表明机场禁止转机
        # Nighttime_ban - 有无宵禁
        # Noise_restrictions - 有无噪声管制
        # Passengers - 客运需求条
        # Cargo - 货运需求条
        t_sql = sqlite3.connect(self.DBPath)
        t_sql.execute(create_sql)
        t_sql.close()

    def DB_TRUNCATE(self):
        t_sql = sqlite3.connect(self.DBPath)
        t_sql.execute("DELETE FROM AirportInfo;")
        t_sql.commit()
        t_sql.close()

    def thread_getAirportInfo(self, AirportNumber: int):
        """
        实际获取机场信息的线程，没有其他的功能
        :param AirportNumber: 航班号码
        """
        # 机场是自增探索形式，如果失败了后边就不需要做了
        t_response = requests.get(self.baseURL + str(AirportNumber), headers=self.get_header)
        if self.errorURL in t_response.url:
            # 这里进行二次路径探测，换句话说就是看后面还有没有，有，说明是单独误报
            flag_found = -1
            print('id为 %d 的机场不存在。' % AirportNumber)
            for i in range(1, 6):
                t_response = requests.get(self.baseURL + str(AirportNumber + max_thread_workers * i),
                                          allow_redirects=False, headers=self.get_header)
                if self.errorURL not in t_response.headers.get('Location', ''):
                    flag_found = AirportNumber + i * 5
                    break
                else:
                    print('id为 %d 的机场不存在。' % (AirportNumber + max_thread_workers * i))
            if flag_found == -1:
                print('线程已退出。')
                self.flag_finish.append((AirportNumber - 1) % max_thread_workers)
            else:
                self.thread_getAirportInfo(flag_found)  # 节省线程使用量
            return
        # 具体执行
        insert_sql = "INSERT INTO AirportInfo VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);"
        t_dict = {"Time_Zone": 0, "IATA_Code": '', "ICAO_Code": '', "Country": '', "Continent": '', "Runway": 0,
                  "Airport_Size": '', "Slots_per_five_minutes": 0, "Slot_Availability": 0, "Min_transfer_time": -1,
                  "Nighttime_ban": 0, "Noise_restrictions": 0, "Passengers": 0, "Cargo": 0, "AirportName": '',
                  "Region": ''}

        def Recursion_ParseAirportInfo(root: bs4_Tag, dataDict: dict):
            if root.name == 'td':
                if root.getText() in ('時區', 'Time zone'):
                    try:
                        dataDict['Time_Zone'] = int(
                            root.parent.contents[1].getText().replace('+', '').replace('UTC', ''). \
                                replace(' ', '').split(':')[0])
                    except:  # 可能遇到了格林尼治时期（UTC时间）
                        pass
                elif root.getText() in ('IATA 代號', 'IATA code'):
                    dataDict['IATA_Code'] = root.parent.contents[1].getText()
                elif root.getText() in ('ICAO 代號', 'ICAO code'):
                    dataDict['ICAO_Code'] = root.parent.contents[1].getText()
                elif root.getText() in ('區域', 'Region'):
                    dataDict['Region'] = root.parent.contents[1].contents[0].getText()
                elif root.getText() in ('國家', 'Country'):
                    dataDict["Country"] = root.parent.contents[1].contents[0].getText()
                elif root.getText() in ('洲別', 'Continent'):
                    dataDict["Continent"] = root.parent.contents[1].getText()
                elif root.getText() in ('跑道', 'Runway'):
                    dataDict["Runway"] = int(
                        root.parent.contents[1].getText().replace(',', '').replace('m', '').strip())
                elif root.getText() in ('機場大小', 'Airport size'):
                    dataDict["Airport_Size"] = root.parent.contents[1].getText()
                # elif root.getText() in ('時間帶數 (每五分鐘)', 'Slots (per 5 minutes)'):
                elif 'Slots (per 5 minutes)' in root.getText() or '時間帶數 (每五分鐘)' in root.getText():
                    dataDict["Slots_per_five_minutes"] = int(root.parent.contents[1].getText())
                elif root.getText() in ('可用時間帶', 'Slot Availability'):
                    dataDict["Slot_Availability"] = int(root.parent.contents[1].getText().replace('%', '')) / 100
                elif root.getText() in ('最短轉機時間', 'Min. transfer time'):
                    t1 = root.parent.contents[1].contents[0].getText()
                    if 'transfer impossible' not in t1 or '不可轉機' not in t1:
                        t1: str = root.parent.contents[1].contents[0].getText()
                        dataDict["Min_transfer_time"] = int(t1.split(':')[0]) * 60 + int(t1.split(':')[1])
                elif root.getText() in ('宵禁', 'Nighttime ban'):
                    if root.parent.contents[1].getText() not in ('無宵禁', 'no nighttime ban'):
                        dataDict["Nighttime_ban"] = 1
                elif root.getText() in ('噪音管制', 'Noise restrictions'):
                    if root.parent.contents[1].getText() not in ('無噪音管制', 'no noise restrictions'):
                        dataDict["Noise_restrictions"] = 1
                elif root.getText() in ('旅客', 'Passengers'):
                    dataDict["Passengers"] = int(root.parent.contents[1].contents[0].attrs.get('title'). \
                                                 replace('demand:', '').strip())
                elif root.getText() in ('貨物', 'Cargo'):
                    dataDict["Cargo"] = int(root.parent.contents[1].contents[0].attrs.get('title'). \
                                            replace('demand:', '').strip())
            elif root.name == 'h1' and ('Airport:' in root.getText() or '機場: ' in root.getText()):
                dataDict["AirportName"] = root.getText().split(':')[1].split('(')[0].strip()
                return

        t_dict.update(CommonHTMLParser(GetClearHTML(t_response), Recursion_ParseAirportInfo))
        t_sql = sqlite3.connect(self.DBPath)
        t_sql.execute(insert_sql, (AirportNumber, t_dict["AirportName"], t_dict["Time_Zone"], t_dict["IATA_Code"],
                                   t_dict["ICAO_Code"], t_dict["Country"], t_dict["Region"], t_dict["Continent"],
                                   t_dict["Runway"], t_dict["Airport_Size"], t_dict["Slots_per_five_minutes"],
                                   t_dict["Slot_Availability"],
                                   t_dict["Min_transfer_time"], t_dict["Nighttime_ban"], t_dict["Noise_restrictions"],
                                   t_dict["Passengers"], t_dict["Cargo"]))
        t_sql.commit()
        t_sql.close()
        print('对机场 %s 信息抓取完成。' % t_dict["IATA_Code"])
        Thread(target=self.thread_getAirportInfo, args=(AirportNumber + max_thread_workers,)).start()

    def DetectFinish(self):
        return len(self.flag_finish) == max_thread_workers

    def DetectAirportUpdate_Basic(self) -> bool:
        """检测机场信息是否有更新，目前比较简单，只获取最新的一个文件名。"""
        LatestFile = ['']

        def Recursion_GetLatestChangelogFileName(root: bs4_Tag):
            if root.name == 'a' and 'airport-data-changelog' in root.attrs.get('href', '').lower():
                LatestFile[0] = root.attrs.get('href')  # 有一定风险，如果日期不是按照从上到下的顺序排列时
            for t_unit in root.children:
                if isinstance(t_unit, bs4_Tag):
                    Recursion_GetLatestChangelogFileName(t_unit)

        for unit in BeautifulSoup(retry_request('https://www.airlinesim.aero/blog/files/').text, 'html5lib'):
            if isinstance(unit, bs4_Tag):
                Recursion_GetLatestChangelogFileName(unit)
        t1 = ConfigManager.LoadConfig(self.DBPath, 'GetAirportInfo')
        if t1.get('LatestAirportInfo', '') != LatestFile[0]:
            t1.update({'LatestAirportInfo': LatestFile[0]})
            ConfigManager.SaveConfig(self.DBPath, 'GetAirportInfo', t1)
            return True
        else:
            return False


class GetBusinessStatistics:
    # 基础信息收集
    AirCompanies = []
    current_TotalPassengerData = {}
    current_TotalCargoData = {}
    current_WeekPassengerData = {}
    current_WeekPassengerCapacityData = {}
    current_WeekCargoData = {}
    current_WeekCargoCapacityData = {}
    map_english_to_chinese = {'ALL': '世界'}
    # 线程池参数（这个进程不需要存储数据库，因此线程将被滥发）
    __pool_workTasks = []  # 工作线程任务池，每个任务参数以tuple(信息分类，大洲，国家)形式
    __pool_hasFinishedThread = []
    # Const
    const_countryName_to_id = {}
    const_continentName_to_id = {}
    const_TotalPassenger = 'entpax'
    const_TotalCargo = 'entcargo'
    const_WeekPassenger = 'transrecpax'
    const_WeekCargo = 'transreccargo'
    const_WeekPassengerCapacity = 'transrecpaxcap'
    const_WeekCargoCapacity = 'transreccargocap'

    def __init__(self, ServerName: str, callback_InitOK=None):
        """
        指定需要提取统计信息的服务器的名称，以及在初始化线程执行完成后调用的回调函数。

        :param ServerName: 服务器名称，请参考LoginAirlineSim.ServerMap
        :param callback_InitOK: 回调函数，无参，将在初始化线程执行完毕前调用
        """
        from LoginAirlineSim import ServerMap, getBaseURL

        if ServerName not in ServerMap.keys():
            raise Exception('必须指定正确的服务器名称！')
        self.baseURL = getBaseURL(ServerName)
        Thread(target=self.__thread_InitBasicInfo).start()
        self.__function_InitOK = callback_InitOK
        self.__group_const_info = (self.const_TotalPassenger, self.const_TotalCargo, self.const_WeekPassenger,
                                   self.const_WeekPassengerCapacity, self.const_WeekCargo, self.const_WeekCargoCapacity)

    def GetBusinessStatistics(self, savePath: str, extractTotalPassenger: bool = False, extractTotalCargo: bool = False,
                              extractWeekPassenger: bool = False, extractWeekPassengerCapacity: bool = False,
                              extractWeekCargo: bool = False, extractWeekCargoCapacity: bool = False,
                              specialContinent: list = None, specialCountries: list = None,
                              specialCompanies: list = None):
        """
        用户级接口，负责该类的对外调用。

        如果指定了企业列表，则从所给的大洲和国家列表之中收集所有企业的信息，并以指定的企业列表筛选出数据。

        如果未指定企业列表，则以大洲和国家列表名列出企业的信息。

        :param savePath: 结果的存放文件，只提供Excel文档。
        :param extractTotalPassenger: 是否导出总旅客运送量的数据。
        :param extractTotalCargo: 是否导出总货物运输量的数据
        :param extractWeekPassenger: 是否导出周旅客运送量的数据。
        :param extractWeekPassengerCapacity: 是否导出周旅客上限承载量的数据。
        :param extractWeekCargo: 是否导出周货物运输量的数据。
        :param extractWeekCargoCapacity: 是否导出周货物上限承载量的数据。
        :param specialContinent: 指定仅收集哪些大洲的信息（单独列出）。
        :param specialCountries: 指定仅收集哪些国家的信息（单独列出）。
        :param specialCompanies: 指定仅收集哪些企业的信息（单独列出），会受到大洲和国家的限制。
        """
        extractFlags = [extractTotalPassenger, extractTotalCargo, extractWeekPassenger, extractWeekPassengerCapacity,
                        extractWeekCargo, extractWeekCargoCapacity]
        assert extractFlags.count(False) + extractFlags.count(True) == 6, '无法识别的一个或多个提取参数。'
        if extractFlags.count(False) == len(extractFlags):
            return
        self.__getBusinessStatisticsData(extractFlags, specialContinent, specialCountries)
        if specialCountries is None and specialContinent is None:
            tableUnitNames = None
        else:
            tableUnitNames = []
            if isinstance(specialContinent, list):
                tableUnitNames += specialContinent
            if isinstance(specialCountries, list):
                tableUnitNames += specialCountries
            if len(tableUnitNames) == 0:
                tableUnitNames = None
        if not savePath.lower().endswith('.xlsx'):
            savePath += '.xlsx'
        self.__outputExcel(savePath, extractFlags, tableUnitNames, specialCompanies)

    def __outputExcel(self, savePath: str, infoExtractPara: list, sheetFilter: list = None,
                      specialCompany: list = None):
        import openpyxl
        cache_agent = [self.current_TotalPassengerData, self.current_TotalCargoData, self.current_WeekPassengerData,
                       self.current_WeekPassengerCapacityData, self.current_WeekCargoData,
                       self.current_WeekCargoCapacityData]

        def addSheetHeader(sheet: openpyxl.workbook.workbook.Worksheet):
            const_first_line = ['总客流量', '总载货量', '周客流量', '周客运承载量', '周载货量', '周载货承载量']
            const_second_line = ['排名', '市场份额', '百分比']
            first_line = ['']
            second_line = ['']
            for infoIndex in range(6):
                if infoExtractPara[infoIndex]:
                    first_line += [const_first_line[infoIndex], '', '']
                    second_line += const_second_line
            sheet.append(first_line)
            sheet.append(second_line)
            for infoIndex in range(infoExtractPara.count(True)):
                sheet.merge_cells('%s1:%s1' % (chr(ord('B') + infoIndex * 3), chr(ord('D') + infoIndex * 3)))

        if not isinstance(sheetFilter, list) or len(sheetFilter) == 0:
            sheetFilter = ['ALL']
        resultWorkbook = openpyxl.Workbook()
        resultWorkbook.remove(resultWorkbook.active)
        if isinstance(specialCompany, list) and len(specialCompany) > 0:
            for company in specialCompany:
                tSheet = resultWorkbook.create_sheet(company)
                addSheetHeader(tSheet)
                for filterStr in sheetFilter:
                    row_data = ['%s/%s' % (self.map_english_to_chinese.get(filterStr, filterStr), filterStr)]
                    for infoID in range(len(infoExtractPara)):
                        if infoExtractPara[infoID] and filterStr in cache_agent[infoID].keys():
                            unit_dict: dict = cache_agent[infoID].get(filterStr, {}).get(company, {})
                            row_data.append(unit_dict.get('Order', 'N/A'))
                            row_data.append(unit_dict.get('MarketShare', '0'))
                            row_data.append(unit_dict.get('Percentage', '0%'))
                    tSheet.append(row_data)
        else:
            for filterStr in sheetFilter:
                tSheet = resultWorkbook.create_sheet('%s/%s' % (
                    self.map_english_to_chinese.get(filterStr, filterStr), filterStr))
                tSet = set()
                for infoID in range(len(infoExtractPara)):
                    if infoExtractPara[infoID]:
                        tSet.update(list(cache_agent[infoID].get(filterStr, {}).keys()))
                for company in list(tSet):
                    row_data = [company]
                    for infoID in range(len(infoExtractPara)):
                        if infoExtractPara[infoID]:
                            unit_dict: dict = cache_agent[infoID].get(filterStr, {}).get(company, {})
                            row_data.append(unit_dict.get('Order', 'N/A'))
                            row_data.append(unit_dict.get('MarketShare', '0'))
                            row_data.append(unit_dict.get('Percentage', '0%'))
                    tSheet.append(row_data)
        try:
            resultWorkbook.save(savePath)
        except:
            from datetime import datetime
            resultWorkbook.save('AS_Total_%s.xlsx' % datetime.now().strftime('%Y%m%d%H%M%S'))

    def __getBusinessStatisticsData(self, infoExtractPara: list, specialContinent: list = None,
                                    specialCountry: list = None):
        """
        根据筛选器构造URL并监控执行，此处需要长时间等待。

        :param infoExtractPara: 提取参数，与group_const_info定义相同，但全是Boolean值
        :param specialContinent: 特定大洲
        :param specialCountry: 特定国家
        """
        if not isinstance(infoExtractPara, list) or len(infoExtractPara) != 6 or infoExtractPara.count(True) < 1 or \
                infoExtractPara.count(True) + infoExtractPara.count(False) != 6:
            raise IndexError('参数infoExtractPara不符合格式！')
        for boolID in range(6):
            if infoExtractPara[boolID]:
                if specialCountry is None and specialContinent is None or (
                        isinstance(specialContinent, list) and len(specialContinent) == 0 and
                        isinstance(specialCountry, list) and len(specialCountry) == 0):
                    if 0 <= boolID <= 1:
                        continue  # 之前已经采集过了，使用缓存信息即可
                    self.__pool_workTasks.append((self.__group_const_info[boolID], 'ALL', 'ALL'))
                else:
                    if isinstance(specialCountry, list):
                        for unit in specialCountry:
                            if isinstance(unit, str) and len(unit) > 0:
                                self.__pool_workTasks.append((self.__group_const_info[boolID], unit, None))
                    if isinstance(specialContinent, list):
                        for unit in specialContinent:
                            if isinstance(unit, str) and len(unit) > 0:
                                self.__pool_workTasks.append((self.__group_const_info[boolID], None, unit))
        total_threads = len(self.__pool_workTasks)
        for threadInt in range(total_threads):
            Thread(target=self.__thread_WorkThread).start()
        while len(self.__pool_hasFinishedThread) < total_threads:
            from time import sleep
            sleep(5)
        self.__pool_hasFinishedThread.clear()

    """
    设计目的：一个可以自动化收集信息并输出为Excel或者HTML（待设计）的报告结果。
    初始可收集信息：企业 - （客运总量/entpax，货运总量/entcargo，周客运量/transrecpax，周客运容量/transrecpaxcap，
                        周货运量/transreccargo，周货运容量/transreccargocap）
    可用条件筛选器：国家、大洲、无分类（世界）
    --------------------------------
    操作流程：信息收集 -> 分类存放 -> 整理输出
    --------------------------------
    基类实现：
    收集特定企业的特定数据
    --------------------------------
    导出笔记：
    优先筛选企业，对空缺部分以“无数据”填补。
    当没有特定企业的时候，以大洲/国家列出存储的企业信息（ALL也算）。
    纵轴为企业名称，横轴为数据。
    """

    def __getInfoURL(self, infoType: str, country: str = None, continent: str = None, unlimited: bool = True):
        """
        构造URL。

        :param infoType: 要收集的信息类型
        :param country: 可选，国家
        :param continent: 可选，大洲
        :param unlimited: 无限制地取数
        :return: 构建好的URL
        """
        result = self.baseURL + '/action/info/stat?type=' + infoType
        if isinstance(country, str) and country in self.const_countryName_to_id.keys():
            result += '&country=' + self.const_countryName_to_id.get(country)
        elif isinstance(continent, str) and continent in self.const_continentName_to_id.keys():
            result += '&continent=' + self.const_continentName_to_id.get(continent)
        if unlimited:
            result += '&limit=0'
        return result

    def __thread_InitBasicInfo(self):
        """初始化基础信息收集，但同时也会收集全局总旅客量信息和全局总货物运输信息。"""
        from PublicCode import Continent_UI, Countries_UI
        for tStr in Continent_UI.keys():
            self.map_english_to_chinese[Continent_UI.get(tStr)] = tStr
        for tStr in Countries_UI.keys():
            self.map_english_to_chinese[Countries_UI.get(tStr)] = tStr
        # 下面开始原操作流程
        startPage = GetClearHTML(retry_request(self.__getInfoURL(self.const_TotalPassenger)))

        def parseHTML_GetCountriesIndex(root: bs4_Tag, dataDict: dict):
            if root.name == 'select' and root.attrs.get('name', '') == 'country':
                for country in root.children:
                    if isinstance(country, bs4_Tag) and country.name == 'option':
                        dataDict[country.getText().strip()] = country.attrs.get('value')
                return True

        def parseHTML_GetContinentIndex(root: bs4_Tag, dataDict: dict):
            if root.name == 'select' and root.attrs.get('name', '') == 'continent':
                for country in root.children:
                    if isinstance(country, bs4_Tag) and country.name == 'option':
                        dataDict[country.getText().strip()] = country.attrs.get('value')
                return True

        self.const_countryName_to_id.update(CommonHTMLParser(startPage, parseHTML_GetCountriesIndex))
        self.const_continentName_to_id.update(CommonHTMLParser(startPage, parseHTML_GetContinentIndex))
        self.current_TotalPassengerData['ALL'] = CommonHTMLParser(startPage, self.parseHTML_GetStatisticsInfo)
        self.current_TotalCargoData['ALL'] = CommonHTMLParser(GetClearHTML(retry_request(self.__getInfoURL(
            self.const_TotalCargo))), self.parseHTML_GetStatisticsInfo)
        tSet = set(list(self.current_TotalPassengerData.keys()))
        tSet.update(list(self.current_TotalCargoData.keys()))
        self.AirCompanies = list(tSet)
        if callable(self.__function_InitOK):
            try:
                self.__function_InitOK()
            finally:
                pass

    def __thread_WorkThread(self):
        try:
            thisTaskPara = self.__pool_workTasks.pop()
            if thisTaskPara[0] not in self.__group_const_info or len(thisTaskPara) < 3:
                self.__pool_hasFinishedThread.append(False)
                return
        except IndexError:
            self.__pool_hasFinishedThread.append(False)
            return
        url = self.__getInfoURL(thisTaskPara[0], thisTaskPara[1], thisTaskPara[2])
        result = CommonHTMLParser(GetClearHTML(retry_request(url)), self.parseHTML_GetStatisticsInfo)
        if isinstance(thisTaskPara[1], str):
            taskKey = thisTaskPara[1]
        else:
            taskKey = thisTaskPara[2]
        if thisTaskPara[0] == self.const_TotalPassenger:
            self.current_TotalPassengerData[taskKey] = result
        elif thisTaskPara[0] == self.const_TotalCargo:
            self.current_TotalCargoData[taskKey] = result
        elif thisTaskPara[0] == self.const_WeekPassenger:
            self.current_WeekPassengerData[taskKey] = result
        elif thisTaskPara[0] == self.const_WeekPassengerCapacity:
            self.current_WeekPassengerCapacityData[taskKey] = result
        elif thisTaskPara[0] == self.const_WeekCargo:
            self.current_WeekCargoData[taskKey] = result
        else:
            self.current_WeekCargoCapacityData[taskKey] = result
        self.__pool_hasFinishedThread.append(True)

    @staticmethod
    def parseHTML_GetStatisticsInfo(root: bs4_Tag, dataDict: dict):
        """仅用于获取统计表单信息，被获取的信息包括：排名、企业名称（Key）、旅客量/货物量、占比数值"""
        if root.name == 'tr' and root.parent.name == 'tbody':
            # 进入了一个行
            order_int = int(root.contents[0].getText())
            company_name = root.contents[1].contents[0].getText()
            tVar = root.contents[-1].getText().strip().split()
            total_int = int(tVar[0].replace(',', ''))
            marketShare_float = tVar[-1]
            dataDict[company_name] = {'Order': order_int, 'MarketShare': total_int, 'Percentage': marketShare_float}
            return True
