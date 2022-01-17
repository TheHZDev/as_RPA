from threading import Thread

import openpyxl
import wx
from openpyxl.workbook.workbook import Workbook, Worksheet
from requests import Session

from GUI_LoginAS import LoginAirlineSimDialog
from NewFlightPlanningSystem import NewFlightPlanningSystem
from PublicCode import TranslateCHTtoCHS as Translate
from PublicCode import openpyxl_ConfigAlignment, Public_ConfigDB_Path


class GUIAutoFlightPlanningBaseOnExcel(wx.Frame):
    generatedSearchOption = None
    nowSearchOption = [False] * 3
    # Const定义块
    const_GenerateFlightInfo = '检索排程系统'
    const_OutputTemplate = '导出排程模板'
    # 标志定义块
    flag_SearchingInfo = False

    def __init__(self, logonSession, serverName):
        wx.Frame.__init__(self, None, id=wx.ID_ANY, title=u"基于Excel的自动排班管理器（AirlineSim专用）", pos=wx.DefaultPosition,
                          size=wx.Size(516, 519),
                          style=wx.CAPTION | wx.CLOSE_BOX | wx.MINIMIZE | wx.MINIMIZE_BOX | wx.TAB_TRAVERSAL)

        self.SetSizeHints(wx.DefaultSize, wx.DefaultSize)
        self.SetBackgroundColour(wx.Colour(207, 243, 216))

        bSizer9 = wx.BoxSizer(wx.VERTICAL)

        gSizer13 = wx.GridSizer(0, 2, 0, 0)

        sbSizer6 = wx.StaticBoxSizer(wx.StaticBox(self, wx.ID_ANY, u"操作选单"), wx.VERTICAL)

        self.GenerateExcelTemplateButton = wx.Button(sbSizer6.GetStaticBox(), wx.ID_ANY, self.const_GenerateFlightInfo,
                                                     wx.DefaultPosition, wx.DefaultSize, 0)
        self.GenerateExcelTemplateButton.SetFont(
            wx.Font(15, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString))
        self.GenerateExcelTemplateButton.SetToolTip(u"检索信息或导出表格模板。")

        sbSizer6.Add(self.GenerateExcelTemplateButton, 0, wx.ALL | wx.EXPAND, 5)

        self.m_staticText281 = wx.StaticText(sbSizer6.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                             wx.Size(-1, 3), 0)
        self.m_staticText281.Wrap(-1)

        sbSizer6.Add(self.m_staticText281, 0, wx.ALL, 5)

        self.InputFlightPlanExcelButton = wx.Button(sbSizer6.GetStaticBox(), wx.ID_ANY, u"导入排程文档", wx.DefaultPosition,
                                                    wx.DefaultSize, 0)
        self.InputFlightPlanExcelButton.SetFont(
            wx.Font(15, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString))
        self.InputFlightPlanExcelButton.Enable(False)
        self.InputFlightPlanExcelButton.SetToolTip(u"导入已填写的表格模板（会自动无视无效航机）。")

        sbSizer6.Add(self.InputFlightPlanExcelButton, 0, wx.ALL | wx.EXPAND, 5)

        self.m_staticText28 = wx.StaticText(sbSizer6.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                            wx.Size(-1, 3), 0)
        self.m_staticText28.Wrap(-1)

        sbSizer6.Add(self.m_staticText28, 0, wx.ALL, 5)

        self.ExecuteInputtedFlightButton = wx.Button(sbSizer6.GetStaticBox(), wx.ID_ANY, u"执行已导入排程", wx.DefaultPosition,
                                                     wx.DefaultSize, 0)
        self.ExecuteInputtedFlightButton.SetFont(
            wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString))
        self.ExecuteInputtedFlightButton.Enable(False)
        self.ExecuteInputtedFlightButton.SetToolTip(u"执行已经导入成功的排程。")

        sbSizer6.Add(self.ExecuteInputtedFlightButton, 0, wx.ALL | wx.EXPAND, 5)

        self.m_staticText282 = wx.StaticText(sbSizer6.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                             wx.Size(-1, 3), 0)
        self.m_staticText282.Wrap(-1)

        sbSizer6.Add(self.m_staticText282, 0, wx.ALL, 5)

        self.ExitButton = wx.Button(sbSizer6.GetStaticBox(), wx.ID_ANY, u"退出", wx.DefaultPosition, wx.DefaultSize, 0)
        self.ExitButton.SetFont(
            wx.Font(15, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString))
        self.ExitButton.SetToolTip(u"退出。")

        sbSizer6.Add(self.ExitButton, 0, wx.ALL | wx.EXPAND, 5)

        gSizer13.Add(sbSizer6, 1, wx.EXPAND, 5)

        sbSizer8 = wx.StaticBoxSizer(wx.StaticBox(self, wx.ID_ANY, u"系统选项"), wx.VERTICAL)

        self.IsSearchSubCompany = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY, u"检索所有子公司", wx.DefaultPosition,
                                              wx.DefaultSize, 0)
        self.IsSearchSubCompany.SetToolTip(u"选中后，检索阶段将搜索全部公司的待排班航班。")

        sbSizer8.Add(self.IsSearchSubCompany, 0, wx.ALL, 5)

        self.IsSearchYellowFlight = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY, u"检索黄色（未执行）排程", wx.DefaultPosition,
                                                wx.DefaultSize, 0)
        sbSizer8.Add(self.IsSearchYellowFlight, 0, wx.ALL, 5)

        self.IsSearchRedFlight = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY, u"检索红色（出现错误）排程", wx.DefaultPosition,
                                             wx.DefaultSize, 0)
        sbSizer8.Add(self.IsSearchRedFlight, 0, wx.ALL, 5)

        self.IsAutoSetUpNewStations = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY, u"自动开设新航站", wx.DefaultPosition,
                                                  wx.DefaultSize, 0)
        self.IsAutoSetUpNewStations.SetToolTip(u"是否自动开设原先并未开设航站的机场和航站。")

        sbSizer8.Add(self.IsAutoSetUpNewStations, 0, wx.ALL, 5)

        self.IsAutoCleanUselessAirlineCode = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY, u"自动清理多余无用航班号码",
                                                         wx.DefaultPosition, wx.DefaultSize, 0)
        self.IsAutoCleanUselessAirlineCode.SetToolTip(u"是否自动清理无用航班号码。")

        sbSizer8.Add(self.IsAutoCleanUselessAirlineCode, 0, wx.ALL, 5)

        gSizer14 = wx.GridSizer(0, 3, 0, 0)

        self.IsStopFlightPlanAfterLowMaintenanceRadio = wx.CheckBox(sbSizer8.GetStaticBox(), wx.ID_ANY,
                                                                    u"当维护比低于\n时停止排班", wx.DefaultPosition,
                                                                    wx.DefaultSize, 0)
        self.IsStopFlightPlanAfterLowMaintenanceRadio.SetToolTip(u"当维护比较低，停止排程。")

        gSizer14.Add(self.IsStopFlightPlanAfterLowMaintenanceRadio, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)

        self.InputMinMaintenanceRadio = wx.TextCtrl(sbSizer8.GetStaticBox(), wx.ID_ANY, u"100", wx.DefaultPosition,
                                                    wx.Size(40, -1), wx.TE_CENTER)
        self.InputMinMaintenanceRadio.Disable()
        gSizer14.Add(self.InputMinMaintenanceRadio, 0, wx.ALL | wx.ALIGN_RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)

        self.m_staticText26 = wx.StaticText(sbSizer8.GetStaticBox(), wx.ID_ANY, u"%", wx.DefaultPosition,
                                            wx.DefaultSize, 0)
        self.m_staticText26.Wrap(-1)

        gSizer14.Add(self.m_staticText26, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)

        sbSizer8.Add(gSizer14, 1, wx.EXPAND, 5)

        gSizer15 = wx.GridSizer(0, 2, 0, 0)

        self.m_staticText27 = wx.StaticText(sbSizer8.GetStaticBox(), wx.ID_ANY, u"航机排班后默认操作", wx.DefaultPosition,
                                            wx.DefaultSize, 0)
        self.m_staticText27.Wrap(-1)

        gSizer15.Add(self.m_staticText27, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL | wx.ALIGN_CENTER_HORIZONTAL, 5)

        DefaultCommitAfterFlightPlanChoiceChoices = [u"无操作", u"立即执行", u"延后三天", u"锁定航班"]
        self.DefaultCommitAfterFlightPlanChoice = wx.Choice(sbSizer8.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition,
                                                            wx.DefaultSize, DefaultCommitAfterFlightPlanChoiceChoices,
                                                            0)
        self.DefaultCommitAfterFlightPlanChoice.SetSelection(0)
        gSizer15.Add(self.DefaultCommitAfterFlightPlanChoice, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)

        sbSizer8.Add(gSizer15, 1, wx.EXPAND, 5)

        gSizer13.Add(sbSizer8, 1, wx.EXPAND, 5)

        bSizer9.Add(gSizer13, 1, wx.EXPAND, 5)

        sbSizer9 = wx.StaticBoxSizer(wx.StaticBox(self, wx.ID_ANY, u"日志输出"), wx.VERTICAL)

        self.UILogOutputText = wx.TextCtrl(sbSizer9.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition,
                                           wx.Size(-1, 195), wx.TE_CHARWRAP | wx.TE_MULTILINE | wx.TE_READONLY)
        sbSizer9.Add(self.UILogOutputText, 0, wx.EXPAND, 5)

        bSizer9.Add(sbSizer9, 1, wx.EXPAND, 5)

        self.SetSizer(bSizer9)
        self.Layout()
        self.m_statusBar2 = self.CreateStatusBar(1, wx.STB_SIZEGRIP, wx.ID_ANY)

        self.Centre(wx.BOTH)
        self.SetStatusBar(self.m_statusBar2)

        # Connect Events
        self.GenerateExcelTemplateButton.Bind(wx.EVT_BUTTON, self.GenerateExcelTemplateButtonOnButtonClick)
        self.InputFlightPlanExcelButton.Bind(wx.EVT_BUTTON, self.InputFlightPlanExcelButtonOnButtonClick)
        self.ExecuteInputtedFlightButton.Bind(wx.EVT_BUTTON, self.ExecuteInputtedFlightButtonOnButtonClick)
        self.ExitButton.Bind(wx.EVT_BUTTON, self.ExitButtonOnButtonClick)
        self.IsSearchSubCompany.Bind(wx.EVT_CHECKBOX, self.IsSearchSubCompanyOnCheckBox)
        self.IsSearchYellowFlight.Bind(wx.EVT_CHECKBOX, self.IsSearchYellowFlightOnCheckBox)
        self.IsSearchRedFlight.Bind(wx.EVT_CHECKBOX, self.IsSearchRedFlightOnCheckBox)
        self.IsStopFlightPlanAfterLowMaintenanceRadio.Bind(wx.EVT_CHECKBOX,
                                                           self.IsStopFlightPlanAfterLowMaintenanceRadioOnCheckBox)
        self.InputMinMaintenanceRadio.Bind(wx.EVT_KILL_FOCUS, self.InputMinMaintenanceRadioOnKillFocus)

        # Init
        self.flightPlanningSystem = FlightPlanningSystemBaseOnExcel(logonSession, serverName, self.callback_PostError,
                                                                    self.callback_LogOutput,
                                                                    self.callback_ShowMessageDialog)

    def GenerateExcelTemplateButtonOnButtonClick(self, event):
        if self.GenerateExcelTemplateButton.GetLabel() == self.const_OutputTemplate:
            from datetime import datetime
            filename = datetime.now().strftime('%Y%m%d') + '.xlsx'
            outputFileDialog = wx.FileDialog(self, '导出模板文件', defaultFile=filename, wildcard='Excel 文档|*.xlsx',
                                             style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
            outputFileDialog.ShowModal()
            savePath: str = outputFileDialog.GetPath()
            outputFileDialog.Destroy()
            if savePath is None or not savePath.endswith('.xlsx'):
                wx.MessageDialog(self, '没有文件被导出。', '信息', wx.ICON_INFORMATION | wx.OK).ShowModal()
                return
            self.SetStatusText('正在导出模板。。。')
            self.GenerateExcelTemplateButton.Disable()
            Thread(target=self.flightPlanningSystem.GenerateExcelTemplateAndOutput,
                   args=(savePath, self.callback_ReEnableUI_AfterGenerateTemplate)).start()
        elif self.GenerateExcelTemplateButton.GetLabel() == self.const_GenerateFlightInfo:
            self.UILogOutputText.Clear()
            self.SetStatusText('正在检索航班信息，请稍等。。。')
            # 状态快照处理
            self.generatedSearchOption = self.nowSearchOption.copy()
            self.GenerateExcelTemplateButton.Disable()
            self.IsSearchSubCompany.Disable()
            self.IsSearchRedFlight.Disable()
            self.IsSearchYellowFlight.Disable()
            self.InputFlightPlanExcelButton.Disable()
            self.ExecuteInputtedFlightButton.Disable()
            Thread(target=self.thread_CollectAirlineInfo, daemon=True).start()

    def InputFlightPlanExcelButtonOnButtonClick(self, event):
        inputFileDialog = wx.FileDialog(self, '导入排程文件', wildcard='Excel 文档|*.xlsx',
                                        style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
        inputFileDialog.ShowModal()
        try:
            self.SetStatusText('开始导入')
            Thread(target=self.flightPlanningSystem.ReadExcelAndBuildConfig, daemon=True,
                   args=(openpyxl.open(inputFileDialog.GetPath(), read_only=True),
                         self.callback_ReEnableUI_AfterReadExcel)).start()
            self.GenerateExcelTemplateButton.Disable()
            self.InputFlightPlanExcelButton.Disable()
            self.ExecuteInputtedFlightButton.Disable()
            self.IsSearchYellowFlight.Disable()
            self.IsSearchSubCompany.Disable()
            self.IsSearchRedFlight.Disable()
        except:
            wx.MessageDialog(self, '打开Excel文档出错！', '错误！', wx.ICON_ERROR | wx.OK).ShowModal()
        finally:
            inputFileDialog.Destroy()

    def ExecuteInputtedFlightButtonOnButtonClick(self, event):
        if wx.MessageDialog(self, '是否立刻启动自动排程？', '最后确认', wx.YES_NO | wx.ICON_QUESTION).ShowModal() == wx.ID_YES:
            self.GenerateExcelTemplateButton.Disable()
            self.InputFlightPlanExcelButton.Disable()
            self.ExecuteInputtedFlightButton.Disable()
            self.IsSearchYellowFlight.Disable()
            self.IsSearchSubCompany.Disable()
            self.IsSearchRedFlight.Disable()
            self.IsStopFlightPlanAfterLowMaintenanceRadio.Disable()
            self.IsAutoSetUpNewStations.Disable()
            self.IsAutoCleanUselessAirlineCode.Disable()
            self.InputMinMaintenanceRadio.Disable()
            self.DefaultCommitAfterFlightPlanChoice.Disable()
            MinMaintenanceRadio = None
            if self.IsStopFlightPlanAfterLowMaintenanceRadio.GetValue():
                MinMaintenanceRadio = float(self.InputMinMaintenanceRadio.GetValue())
            Thread(target=self.flightPlanningSystem.Thread_FlightManager, daemon=True,
                   args=(self.IsAutoCleanUselessAirlineCode.GetValue(), self.IsAutoSetUpNewStations.GetValue(),
                         MinMaintenanceRadio, self.DefaultCommitAfterFlightPlanChoice.GetSelection())).start()
            wx.MessageDialog(self, '航机自动排程已经启动，请在日志窗口查看详细日志。', '提示').Show()
        event.Skip()

    def ExitButtonOnButtonClick(self, event):
        self.flightPlanningSystem.PrepareExit()
        self.Hide()
        self.Close(True)
        event.Skip()

    def IsSearchSubCompanyOnCheckBox(self, event):
        self.nowSearchOption[0] = self.IsSearchSubCompany.GetValue()
        self.checkGenerateButton()
        event.Skip()

    def IsSearchYellowFlightOnCheckBox(self, event):
        self.nowSearchOption[1] = self.IsSearchYellowFlight.GetValue()
        self.checkGenerateButton()
        event.Skip()

    def IsSearchRedFlightOnCheckBox(self, event):
        self.nowSearchOption[2] = self.IsSearchRedFlight.GetValue()
        self.checkGenerateButton()
        event.Skip()

    def IsStopFlightPlanAfterLowMaintenanceRadioOnCheckBox(self, event):
        if self.IsStopFlightPlanAfterLowMaintenanceRadio.GetValue():
            self.InputMinMaintenanceRadio.Enable()
        else:
            self.InputMinMaintenanceRadio.Disable()

    def InputMinMaintenanceRadioOnKillFocus(self, event):
        tStr: str = self.InputMinMaintenanceRadio.GetValue()
        if not (tStr.isdigit() or (tStr.count('.') == 1 and tStr.replace('.', '').isdigit())):
            self.InputMinMaintenanceRadio.SetValue('100')
        event.Skip()

    # 回调函数区
    def callback_LogOutput(self, logStr: str):
        from datetime import datetime
        self.UILogOutputText.AppendText(datetime.now().strftime('%H:%M:%S : ') + logStr + '\n')
        self.SetStatusText(logStr)

    def callback_PostError(self, errorText: str):
        self.callback_ShowMessageDialog(errorText, '错误')
        # 弹窗告警 + 文本输出
        self.callback_LogOutput(errorText)

    def callback_ShowMessageDialog(self, msgText: str, titleText: str = '告警'):
        wx.CallAfter(self.thread_ShowMessageDialog, msgText, titleText)

    def callback_ReEnableUI_AfterGenerateTemplate(self):
        self.GenerateExcelTemplateButton.Enable()
        self.SetStatusText('导出完成。')
        self.callback_ShowMessageDialog('导出成功。', '信息')

    def callback_ReEnableUI_AfterReadExcel(self, result: dict):
        self.InputFlightPlanExcelButton.Enable()
        self.GenerateExcelTemplateButton.Enable()
        self.IsSearchSubCompany.Enable()
        self.IsSearchRedFlight.Enable()
        self.IsSearchYellowFlight.Enable()
        self.callback_ShowMessageDialog('排程信息导入成功！', '导入完成')
        self.SetStatusText('导入成功！')
        if len(result) > 0:
            self.ExecuteInputtedFlightButton.Enable()
        else:
            self.ExecuteInputtedFlightButton.Disable()

    def thread_ShowMessageDialog(self, msgText: str, titleText: str):
        wx.MessageDialog(self, msgText, titleText, wx.ICON_INFORMATION | wx.OK).ShowModal()

    # 规程定义区
    def checkGenerateButton(self):
        if isinstance(self.generatedSearchOption, list):
            if self.nowSearchOption == self.generatedSearchOption:
                self.GenerateExcelTemplateButton.SetLabel(self.const_OutputTemplate)
            else:
                self.GenerateExcelTemplateButton.SetLabel(self.const_GenerateFlightInfo)

    def thread_CollectAirlineInfo(self):
        try:
            self.flightPlanningSystem.SearchInfoIntelligently(self.generatedSearchOption[0],
                                                              self.generatedSearchOption[1],
                                                              self.generatedSearchOption[2])
        finally:
            if len(self.flightPlanningSystem.cache_info) > 0:
                self.GenerateExcelTemplateButton.SetLabel(self.const_OutputTemplate)
                self.InputFlightPlanExcelButton.Enable()
                if len(self.flightPlanningSystem.InputtedFlightPlanData) > 0:
                    self.ExecuteInputtedFlightButton.Enable()
                self.SetStatusText('已完成。')
                wx.MessageDialog(self, '检索排程信息完成！', '完成', wx.ICON_INFORMATION | wx.OK).ShowModal()
            self.GenerateExcelTemplateButton.Enable()
            self.IsSearchSubCompany.Enable()
            self.IsSearchRedFlight.Enable()
            self.IsSearchYellowFlight.Enable()


class FlightPlanningSystemBaseOnExcel(NewFlightPlanningSystem):
    InputtedFlightPlanData = {}

    __flag_ExitProgram = False

    def __init__(self, LogonSession: Session, ServerName: str, callback_ReportError, callback_ShowProgressText,
                 callback_ShowInstantMsg):

        super().__init__(LogonSession, ServerName, callback_ReportError=callback_ReportError,
                         callback_ShowProgressText=callback_ShowProgressText)

        self.function_ShowInstantMsg = callback_ShowInstantMsg

    def PrepareExit(self):
        self.__flag_ExitProgram = True

    def GenerateExcelTemplateAndOutput(self, SavePath: str, Callback_ReEnableUI=None):
        """
        导出排班模板数据并生成Excel文档。

        :param SavePath: 模板文件的保存路径
        :param Callback_ReEnableUI: 重启UI的回调函数
        """
        if len(self.cache_info) < 1:
            self.basic_ReportError('没有收集任何信息，无需导出。')
            return
        outputWorkbook = openpyxl.Workbook()
        self.__initHelpTextInExcel(outputWorkbook)
        self.__initStationsInfoInExcel(outputWorkbook)
        self.__initFlightTemplateInExcel(outputWorkbook)
        try:
            outputWorkbook.save(SavePath)
        except:
            from datetime import datetime
            SavePath = datetime.now().strftime('%Y%m%d_%H%M%S.xlsx')
            outputWorkbook.save(SavePath)
        finally:
            self.basic_ShowProgress('排班模板已保存到%s中。' % SavePath)
            outputWorkbook.close()
        if callable(Callback_ReEnableUI):
            Callback_ReEnableUI()

    def __initHelpTextInExcel(self, newWorkbook: openpyxl.Workbook):
        """初始化帮助信息，并将服务方案示例写入当前表格"""
        help_Text = [('列名', '帮助说明'), ('所属公司', '此航机所属的企业的名称。'), ('航机型号', '此航机的航机型号。'), ('航机编号', '此航机在游戏中的唯一注册编号。'),
                     ('航机健康度', '此航机的当前健康度。'), ('航机维护比', '此航机在排程之前的预期维护比。'), ('航机机龄', '此航机的服役年龄。'),
                     ('航机任务', '航机当前的任务，停放或处于飞行中。'), ('位置', '此航机目前的机场。'), ('排程状态', '当前航机的排程状态，正常或未执行或出现错误。'), ('', ''),
                     ('出发时间', '航机的出发时间，一般只需要指定第一班航班的出发时间即可。'),
                     ('出发机场', '航机的出发机场，第一班不填则默认为到达机场（若航班正在飞行）或当前待机机场。后续不填则视为前一班的目的机场。'),
                     ('出发航站楼', '出发的航站楼，请根据实际情况处理，不填或填错默认为T1。'), ('目的机场', '航机的目的机场。'),
                     ('目的航站楼', '到达的航站楼，请根据实际情况处理，不填或填错默认为T1。'), ('价格系数', '价格系数，请填写数字[50, 200]。'),
                     ('服务方案', '服务方案，请在下面的表单里寻找内容并确定填写一致，否则程序无法识别。'),
                     ('加速/减速', '仅支持最高速和最低速，不填则为游戏推荐的最佳巡航速度。请填写“加速”，“减速”或“常速”（可以不填）。'),
                     ('周计划排班', '请根据实际需要填写，使用1234567来区分周一到周日。例如1237代表飞周一周二周三周日，不飞周四周五周六。'),
                     ('备注', '可以随意填写，程序不会处理这部分的内容。')]
        first_sheet = newWorkbook.active
        first_sheet.title = '帮助文档'
        first_sheet.column_dimensions['A'].width = (max([len(i[0]) for i in help_Text]) + 1) * 2
        first_sheet.column_dimensions['B'].width = (max([len(i[1]) for i in help_Text]) + 1) * 2
        for line in help_Text:
            first_sheet.append(line)
        # 初始化帮助文档后，初始化服务方案
        first_sheet.append(('',))  # 空一行
        service_list = []
        for subCompany in self.cache_info.keys():
            if 'Service' in self.cache_info.get(subCompany).keys():
                service_list.append(['', '', subCompany] + self.cache_info.get(subCompany).get('Service'))
        first_sheet.append(('', '', '企业名', '服务方案参考（每格为一个方案名）'))
        merge_width = max([len(i) - 3 for i in service_list])
        if merge_width > 1:
            first_sheet.merge_cells('D23:%s23' % chr(67 + merge_width))
        for columnID in range(merge_width + 1):
            cache_column = []
            for sub_service_list in service_list:
                if columnID + 2 < len(sub_service_list):
                    cache_column.append(len(sub_service_list[columnID + 2]))
            first_sheet.column_dimensions[chr(67 + columnID)].width = max(cache_column) + 2
        for line in service_list:
            first_sheet.append(line)

    def __initStationsInfoInExcel(self, newWorkBook: Workbook):
        """在排班模板中插入子公司的航站信息表"""
        for SubCompany_name in self.cache_info.keys():
            if 'StationsInfo' not in self.cache_info.get(SubCompany_name).keys():
                continue
            station_table = newWorkBook.create_sheet(SubCompany_name)
            tableHead1 = [''] * 5 + ['客运'] + [''] * 3 + ['货运']
            station_table.append(tableHead1)
            station_table.merge_cells('F1:I1')
            station_table.merge_cells('J1:M1')
            # 顶部表头处理
            tableHead2 = ('IATA代码', '机场名称', '机场所在地', '有无噪音管制', '有无宵禁',
                          '客运需求条（Bar）', '客运航厦容量', '自有设施处理量', '酬载统计',
                          '货运需求条（Bar）', '货运航厦容量', '自有设施处理量', '酬载统计',
                          '额外可用客运航站楼')
            station_table.append(tableHead2)
            # 细分表头处理
            readonly_StationsInfo: dict = self.cache_info.get(SubCompany_name).get('StationsInfo')
            all_line = []
            for IATA_Code in readonly_StationsInfo.keys():
                readonly_Station: dict = readonly_StationsInfo.get(IATA_Code)
                lineData = [IATA_Code, readonly_Station.get('Name'), Translate(readonly_Station.get('Country')),
                            ('有', '无')[(True, False).index(readonly_Station.get('IsNoiseControl'))],
                            ('有', '无')[(True, False).index(readonly_Station.get('IsCurfew'))]]
                lineData += readonly_Station.get('Passengers') + readonly_Station.get('Cargo') + ['、'.join(
                    readonly_Station.get('ExtraTerminal', []))]
                all_line.append(lineData)
            # 缓存建立完成
            column_Length = [10, max([len(i[1]) for i in all_line]), max(12, max([len(i[2]) for i in all_line])), 14,
                             10, 20, max(14, max([len(i[6]) for i in all_line])),
                             max(16, max([len(i[7]) for i in all_line])), 10, 20,
                             max(14, max([len(i[10]) for i in all_line])), max(16, max([len(i[11]) for i in all_line])),
                             max(20, max([len(i[12]) for i in all_line]))]
            for columnID in range(len(column_Length)):
                station_table.column_dimensions[chr(65 + columnID)].width = column_Length[columnID]
            # 调整列宽
            for line in all_line:
                station_table.append(line)
            openpyxl_ConfigAlignment(station_table, 'A1:N%d' % (len(all_line) + 2), 'center')
        # 完成对航站信息的序列化建表

    def __initFlightTemplateInExcel(self, newWorkBook: Workbook):
        """按公司填写待排班的航机信息，但最后的工作表的表名是航机注册号"""
        pre_table_head_1 = ('所属公司', '航机型号', '航机健康度', '航机维护比', '航机机龄', '航机任务', '位置', 'Y/C/F', '排程状态')
        pre_table_head_2 = ('出发时间', '出发机场', '出发航站楼', '目的机场', '目的航站楼', '价格系数', '服务方案', '加速/减速', '周计划排班', '备注')
        for subCompany in self.cache_info.keys():
            subFleets: dict = self.cache_info.get(subCompany).get('Fleets', {})
            for sheet_name in subFleets.keys():
                airplane_unit: dict = subFleets.get(sheet_name)
                line_data = [subCompany, airplane_unit.get('AirType'), str(airplane_unit.get('Health')) + '%',
                             str(airplane_unit.get('MaintenanceRadio')) + '%', str(airplane_unit.get('Age')) + '年',
                             airplane_unit.get('CurrentTask'), airplane_unit.get('Location'),
                             '/'.join([str(i) for i in airplane_unit.get('Y/C/F')])]
                if isinstance(airplane_unit.get('Location'), tuple):
                    line_data[6] = airplane_unit.get('Location')[0]
                else:
                    line_data[6] = airplane_unit.get('Location')
                if not airplane_unit.get('IsNeedInit'):
                    line_data.append('正常（待排班）')
                elif airplane_unit.get('Yellow/Red'):
                    line_data.append('已排程（未执行）')
                else:
                    line_data.append('已排程（错误）')
                new_table = newWorkBook.create_sheet(sheet_name)
                new_table.append(pre_table_head_1)
                new_table.append(line_data)
                new_table['I2'].hyperlink = airplane_unit.get('url')
                new_table.append(pre_table_head_2)
                # 调整列宽
                openpyxl_ConfigAlignment(new_table, 'A1:J100', 'center')
                column_length = [max(10, len(line_data[0])) + 2, max(10, len(line_data[1])) + 2, 12, 12, 12, 10, 10, 13,
                                 20, 40]
                column_range = list("ABCDEFGHIJ")
                for columnID in range(len(column_range)):
                    new_table.column_dimensions[column_range[columnID]].width = column_length[columnID]

    def ReadExcelAndBuildConfig(self, openedWorkBook: Workbook, Callback_ReEnableUI=None):
        """
        内部工作流程：
        1、读取缓存字典数据，确定各个公司的待排班航机列表、服务方案和航站数据。
        2、按公司的航机注册号序列，首先查找该注册号是否在Excel的工作表中注册，有，则继续下一步，否，则继续比对下一个。
        3、打开已存在序列号同名的工作表，检查D4（第一行的目的机场）是否有数据，没有则直接关闭工作表。
        4、逐行读取数据并进行数据匹配。
            - 如果该行没有目的机场，工作表将关闭，所有数据保存
            - 如果该行有时间（A列），则读取时间数据并进行校验（00：00 至 23：59），合规后存档（默认值-1，-1）
            - 如果该行有出发机场（B列），读取该出发机场，直接存档
                - 如果未检测到有出发机场，第一行的源机场为飞机当前停放地或调机目的地
                - 第二行及以后的出发机场使用前一行的目的机场
            - 如果该行有出发航站楼（C列），校验航站楼是否存在（T1不需要校验），合规后存档
            - 读取目的机场（D列），直接存档
            - 如果该行有出发航站楼（E列），校验航站楼是否存在（T1不需要校验），合规后存档
            - 检测该行的价格系数（F列），有，进行数字合规性校验（50 ~ 200），合规后存档
                - 未检测到，使用上一行数据
                - 第一行未指定价格数据，使用默认值100
            - 检测该行的服务方案，与企业的服务方案缓存进行对比，未比中，则采用服务方案列表里最后一个方案作为默认
                - 未检测到，使用上一行的服务方案（第一行使用最后一个服务方案）
            - 检测该行的加速方案，若有数值则比对预设方案，合规后存档。
                - 未检测到，默认为常速
            - 检测该行的周计划，若检测到1234567以外的字符，直接丢掉该字符，剩下的存档，若最后没有剩下任何东西，数据不存档
        """
        cache_readonly_AirCompanyInfo = {}
        result = {}
        for airCompany in self.cache_info.keys():
            if 'Fleets' in self.cache_info.get(airCompany).keys():
                line_dict = {'FleetIndex': self.cache_info.get(airCompany).get('Fleets').keys(),
                             'Stations': self.cache_info.get(airCompany).get('StationsInfo').copy(),
                             'Service': self.cache_info.get(airCompany).get('Service').copy()}
                cache_readonly_AirCompanyInfo.update({airCompany: line_dict})
        # 建立只读缓存信息拷贝，没有检测到航机的企业默认不需要排班，直接跳过
        for airCompany in cache_readonly_AirCompanyInfo.keys():
            currentCompany: dict = cache_readonly_AirCompanyInfo.get(airCompany)
            stationsInfo: dict = currentCompany.get('Stations')
            serviceList: list = currentCompany.get('Service')
            cache_sheetsName: list = openedWorkBook.sheetnames
            cache_airlinesInfo = {}
            for registerID in currentCompany.get('FleetIndex'):
                if registerID in cache_sheetsName:
                    currentSheet: Worksheet = openedWorkBook[registerID]
                    if currentSheet['D4'].value is None or len(str(currentSheet['D4'].value).strip()) < 1:
                        # 检查失败，该工作表可能是没有删除的模板工作表的一部分
                        continue
                    airlineOrder = []  # 有序存档记录航线
                    rowIndex = 3
                    while True:
                        rowIndex += 1
                        if currentSheet['D%d' % rowIndex].value is None or \
                                str(currentSheet['D%d' % rowIndex].value).strip() == '':
                            break
                        # 目的机场为空，则不继续往下查询
                        airline_dict = {}
                        """
                        内部元素及其含义：
                        DepartureHour - 起飞时间的小时数（可能不存在）
                        DepartureMinute - 起飞时间的分钟数
                        SrcAirport - 出发机场
                        DstAirport - 到达/目的机场
                        SrcTerminal - 出发机场使用的航站楼，T(int)
                        DstTerminal - 到达/目的机场使用的航站楼
                        Price - 价格系数，50 ~ 200
                        Service - 服务方案名称
                        Week - 周排班计划，[bool] * 7
                        Speed - 速度安排，大于0加速，小于0减速
                        """
                        if currentSheet['A%d' % rowIndex].value is not None:
                            departureTime: str = str(currentSheet['A%d' % rowIndex].value).strip()
                            try:
                                departureTimes = departureTime.split(':')
                                departureHour = int(departureTimes[0])
                                departureMinute = int(departureTimes[1])
                                if 0 <= departureHour < 24 and 0 <= departureMinute < 60:
                                    airline_dict.update({'DepartureHour': departureHour,
                                                         'DepartureMinute': departureMinute})
                            except:
                                self.__advance_ReportExcelError(registerID, ('A', rowIndex), '不是一个正确的时间！')
                        # 检测起飞时间 OK
                        if currentSheet['B%d' % rowIndex].value is not None:
                            airline_dict['SrcAirport'] = str(currentSheet['B%d' % rowIndex].value).strip()
                        elif len(airlineOrder) == 0:
                            location: str = self.cache_info.get(airCompany).get('Fleets').get(registerID).get(
                                'Location')
                            if '->' in location:
                                airline_dict['SrcAirport'] = location.split('->')[1]
                            else:
                                airline_dict['SrcAirport'] = location
                        else:
                            airline_dict['SrcAirport'] = airlineOrder[-1].get('DstAirport')
                        # 出发机场 OK
                        if currentSheet['C%d' % rowIndex].value is not None and str(
                                currentSheet['C%d' % rowIndex].value).strip().upper() != 'T1':
                            departureTerminal = str(currentSheet['C%d' % rowIndex].value).strip().upper()
                            if airline_dict.get('SrcAirport') not in stationsInfo.keys() or departureTerminal not in \
                                    stationsInfo.get(airline_dict['SrcAirport']).get('ExtraTerminal', []):
                                # 未找到机场，或者找不到对应的额外航站楼，直接报错
                                self.__advance_ReportExcelError(registerID, ('C', rowIndex), '没发现航站楼信息，已重置为T1。')
                            else:
                                airline_dict['SrcTerminal'] = departureTerminal
                        # 出发航站楼 OK
                        airline_dict['DstAirport'] = str(currentSheet['D%d' % rowIndex].value)
                        # 到达机场 无需检定
                        if currentSheet['E%d' % rowIndex].value is not None and str(
                                currentSheet['E%d' % rowIndex].value).strip().upper() != 'T1':
                            arriveTerminal = str(currentSheet['E%d' % rowIndex].value).strip().upper()
                            if airline_dict.get('DstAirport') not in stationsInfo.keys() or arriveTerminal not in \
                                    stationsInfo.get(airline_dict['DstAirport']).get('ExtraTerminal', []):
                                # 未找到机场，或者找不到对应的额外航站楼，直接报错
                                self.__advance_ReportExcelError(registerID, ('E', rowIndex), '没发现航站楼信息，已重置为T1。')
                            else:
                                airline_dict['DstTerminal'] = arriveTerminal
                        # 到达航站楼 OK
                        if currentSheet['F%d' % rowIndex].value is not None:
                            try:
                                price_value = int(str(currentSheet['F%d' % rowIndex].value).strip())
                                if 50 <= price_value <= 200:
                                    airline_dict['Price'] = price_value
                            except:
                                self.__advance_ReportExcelError(registerID, ('F', rowIndex), '价格系数不正确！')
                        if 'Price' not in airline_dict.keys():
                            if len(airlineOrder) == 0:
                                airline_dict['Price'] = 100
                                self.basic_ReportError('工作表%s发生价格系数未指定异常，已重置为100。' % registerID)
                            else:
                                airline_dict['Price'] = airlineOrder[-1].get('Price')
                        # 价格系数 OK
                        if currentSheet['G%d' % rowIndex].value is not None:
                            service_value = str(currentSheet['G%d' % rowIndex].value).strip()
                            if service_value in serviceList:
                                airline_dict['Service'] = service_value
                        elif len(airlineOrder) == 0:
                            airline_dict['Service'] = serviceList[-1]
                            self.basic_ReportError('工作表%s发生服务方案未指定异常，服务方案已重置为%s。' % (registerID, serviceList[-1]))
                        else:
                            airline_dict['Service'] = airlineOrder[-1].get('Service')
                        # 服务方案 OK
                        if currentSheet['H%d' % rowIndex].value is not None:
                            speed_value = str(currentSheet['H%d' % rowIndex].value).strip()
                            if speed_value == '加速':
                                airline_dict['Speed'] = 1
                            elif speed_value == '减速':
                                airline_dict['Speed'] = -1
                        # 速度 OK
                        if currentSheet['I%d' % rowIndex].value is not None:
                            try:
                                week_value = str(currentSheet['I%d' % rowIndex].value).strip()
                                weekDay = [False] * 7
                                for day in list(week_value):
                                    if day in list('1234567'):
                                        weekDay['1234567'.index(day)] = True
                                if weekDay == [False] * 7:
                                    weekDay = [True] * 7
                                    self.__advance_ReportExcelError(registerID, ('I', rowIndex),
                                                                    '周排班计划不正确！已重置为一周全排。')
                                airline_dict['Week'] = weekDay
                            except:
                                self.__advance_ReportExcelError(registerID, ('I', rowIndex),
                                                                '周排班计划不正确！已重置为一周全排。')
                                airline_dict['Week'] = [True] * 7
                        else:
                            airline_dict['Week'] = airlineOrder[-1].get('Week')
                        # 周计划排班 OK
                        # ALL OK
                        airlineOrder.append(airline_dict)
                    if len(airlineOrder) > 0:
                        cache_airlinesInfo[registerID] = airlineOrder
            if len(cache_airlinesInfo) > 0:
                result[airCompany] = cache_airlinesInfo
        # 最终映射关系：类内缓存器 -> 所属企业 --dict-> 所属航机 --dict-> 排程列表 --list-> 具体排程(dict)
        cache_readonly_AirCompanyInfo.clear()
        self.InputtedFlightPlanData.clear()
        self.InputtedFlightPlanData.update(result)
        if callable(Callback_ReEnableUI):
            Callback_ReEnableUI(result)
        else:
            return result

    def Thread_FlightManager(self, AutoCleanUselessAirlineCode: bool = False, AutoSetUpNewStations: bool = False,
                             MinMaintenanceRadio: float = None, DefaultCommit: int = -1):
        """
        航班管理器排程管理线程，其将会根据所给的排程系统对各个已载入的航班进行排程。

        :param AutoCleanUselessAirlineCode: 是否自动清理无用航班号码
        :param AutoSetUpNewStations: 是否自动开设新航站
        :param MinMaintenanceRadio: 最低维护比例
        :param DefaultCommit: 默认执行的操作
        """
        if AutoCleanUselessAirlineCode:
            self.ClearUnusableAirlineNumber()
        cache_NoStations = self.__analyzeNoExistsStation()
        for company in self.InputtedFlightPlanData.keys():
            self.SwitchToSubCompany(company)
            self.runtime_cache_service.clear()
            self.runtime_cache_stations.clear()
            cache_Fleets: dict = self.cache_info.get(company).get('Fleets')
            if AutoSetUpNewStations and company in cache_NoStations.keys():
                self.SetUpNewStation(cache_NoStations.get(company))
            flag_lowMaintainRadio = -1.0
            for registeredID in self.InputtedFlightPlanData.get(company).keys():
                try:
                    airlineManageURL = cache_Fleets.get(registeredID).get('url')
                    if len(self.runtime_cache_service) == 0 or len(self.runtime_cache_stations) == 0:
                        self.PreCollectRuntimeAirlineBasicInfo(airlineManageURL)
                    if cache_Fleets.get(registeredID).get('IsNeedInit'):
                        self.ExecuteAirlinePlan(4, airlineManageURL)  # 删除排程计划
                    lastResponse = self.retryGET(airlineManageURL)
                    for airline in self.InputtedFlightPlanData.get(company).get(registeredID):
                        result = self.EstablishNewAirline(airlineManageURL, airline.get('SrcAirport'),
                                                          airline.get('DstAirport'), airline.get('Price'),
                                                          airline.get('Service'), airline.get('Week'),
                                                          airline.get('SrcTerminal'), airline.get('DstTerminal'),
                                                          airline.get('DepartureHour', -1),
                                                          airline.get('DepartureMinute', -1),
                                                          airline.get('Speed', 0), lastResponse)
                        lastResponse = result.get('lastResponse')
                        self.basic_ShowProgress('航机 %s 的航线（%s -> %s）已排程。' % (registeredID, airline.get('SrcAirport'),
                                                                             airline.get('DstAirport')))
                        if isinstance(MinMaintenanceRadio, float) or isinstance(MinMaintenanceRadio, int):
                            if result.get('maintenance_ratio', 2 ** 32) <= MinMaintenanceRadio:
                                flag_lowMaintainRadio = lastResponse.get('maintenance_ratio')
                                break
                        if self.__flag_ExitProgram:
                            return
                    if flag_lowMaintainRadio > 1:
                        self.__showInstantMsg('航机 %s 的维护比为 %.2f，排程已终止！' % (registeredID, flag_lowMaintainRadio))
                    elif isinstance(DefaultCommit, int) and 1 <= DefaultCommit <= 4:
                        self.ExecuteAirlinePlan(DefaultCommit, airlineManageURL, lastResponse)
                        self.basic_ShowProgress('航机 %s 的排班已完成，并正常提交。' % registeredID)
                    else:
                        self.basic_ShowProgress('航机 %s 的排班已完成。' % registeredID)
                except Exception as e:
                    self.basic_ReportError('航机 %s 排班失败，尝试下一架中。\n错误代码：%s' % (registeredID, str(e)))
            self.basic_ShowProgress('企业 %s 的航机排程任务已结束。' % company)
        self.__showInstantMsg('全部排程任务已结束！您现在可以退出程序了！')

    def __analyzeNoExistsStation(self):
        """以已收集的航站信息表为参照，解析未开设航站。"""
        resultDict = {}
        for company in self.InputtedFlightPlanData.keys():
            referenceCode = []
            for station in self.cache_info.get(company).get('StationsInfo').keys():
                referenceCode.append(station)
                referenceCode.append(self.cache_info.get(company).get('StationsInfo').get(station).get('Name'))
            resultSet = set()
            for registeredID in self.InputtedFlightPlanData.get(company).keys():
                toAnalyzeList: list = self.InputtedFlightPlanData.get(company).get(registeredID)
                for airline in toAnalyzeList:
                    if airline.get('SrcAirport') not in referenceCode:
                        resultSet.add(airline.get('SrcAirport'))
                    if airline.get('DstAirport') not in referenceCode:
                        resultSet.add(airline.get('DstAirport'))
            if len(resultSet) > 0:
                resultDict[company] = list(resultSet)
        return resultDict

    def __advance_ReportExcelError(self, worksheetName: str, unitName: tuple, errorMsg: str):
        """
        Excel内部事务处理报错模块

        :param worksheetName: Excel工作表名称
        :param unitName: 单元名称（列，行）
        :param errorMsg: 错误文本
        """
        try:
            errorText = '在工作表%s' % worksheetName + '的%s%d发生错误，错误信息为：' % unitName + errorMsg
            self.basic_ReportError(errorText)
        except:
            pass

    def __showInstantMsg(self, msgText: str):
        try:
            self.function_ShowInstantMsg(msgText)
        except:
            pass


def callback_afterLogonInit(logonSession, serverName):
    """对话框后，才初始化主窗口"""
    global mainWin, mainSession
    preDialog.Destroy()
    mainWin = GUIAutoFlightPlanningBaseOnExcel(logonSession, serverName)
    mainSession = logonSession
    mainWin.Show()


if __name__ == '__main__':
    mainAPP = wx.App()
    mainWin = None
    mainSession = None
    try:
        preDialog = LoginAirlineSimDialog(None, Public_ConfigDB_Path, callback_afterLogonInit)
        preDialog.Show()
        mainAPP.MainLoop()
    finally:
        from LoginAirlineSim import LogoutAirlineSim
        from sys import exit

        LogoutAirlineSim(mainSession)
        exit(0)
